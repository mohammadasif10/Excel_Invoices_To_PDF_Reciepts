import pandas as pd#can read csv AND excel
import glob
from fpdf import FPDF
from pathlib import Path#takes a filepath and can do things to the name
import openpyxl#you can just pip install this, inlcuded it here for refrence

filepaths = glob.glob("invoices/*.xlsx")#glob looks for patterns. * is a placeholder. it means anything can be here and since .xlsx is after it. it looping through a folder that has invoices/ at the beginning, anything in the middle, but .xlsx at the end and stores as a list. doesn't have to be .xlsx, can jsut be xlsx. you can place the * anywhere in where you are typing like in this case it's in the middle. just gotta see what pattern you are looking for
print(filepaths)

#this loop is really long and has another for loop inside it too. don't get confused. everything until the pdf get's outputted is inside this loop. outer big loop will itetrate 3 times total in this case bc only 3 files, the inner loop will loop a bunch of times in each file till it writes each row.

for fp in filepaths: 
  #cleans up the filepath variable so we can get just the name of the file
  filename = Path(fp).stem#path from pathlib, takes the filepath and with stem, removes the invoices/ and the .xlsx, we are left w invoice number and date. just splitting would take too long bc splitting first at / and then xlsx. this does it one step, more useful on longer file names
  invNum = filename.split("-")[0]#split stores both halves in a list, we want the first part so we add the [0]
  date = filename.split("-")[1]
  
  #create pdf
  mypdf = FPDF(orientation="P", unit="mm", format="A4")
  mypdf.add_page()
  mypdf.set_font(family="Times", size=16, style="B")
  mypdf.cell(w=50, h=8, txt=f"Invoice Number: {invNum}", ln=1)#ln=1 adds 1 breakline after writing
  mypdf.cell(w=50, h=8, txt=f"Date: {date}", ln=1)

  #read excel file
  excelFile = pd.read_excel(fp, sheet_name="Sheet 1")#panda is reading the excel sheet that's  being looped over rn, have to give a sheet name bc excel can have diff sheets, saved excel file is called a dataframe or any 2d table

  #writing column names onto file
  myColumns = list(excelFile.columns)#just file.columns gives us weird Index data type. we make it into list so easy to work with
  myColumns = [x.replace("_", " ") for x in myColumns]#in line for loop, breaking it down. my columns is a list object. read it backwords. we are looping through each item in myColumns each iteration making the value of x whatever the item of the list we are at. we just replace the _ with a space 
  myColumns = [x.title() for x in myColumns]#in video combined into one loop like x.reaplce("_", " ").title(). i didn't know you can do two functions like that and looked too complicated for out inline for loop understanding so i broke it up into two
  mypdf.set_font(family="Times", size=10, style='B')
  mypdf.set_text_color(80,80,80)
  mypdf.cell(w=30, h=8, txt=myColumns[0], border=1) #getting the items from myColumns list
  mypdf.cell(w=70, h=8, txt=myColumns[1], border=1)
  mypdf.cell(w=30, h=8, txt=myColumns[2], border=1)
  mypdf.cell(w=30, h=8, txt=myColumns[3], border=1)
  mypdf.cell(w=30, h=8, txt=myColumns[4], border=1, ln=1)
  
  #looping over rows and writing excel rows onto pdf
  for index, row in excelFile.iterrows():#remeber, iterrows gives two values which is why there's two variables in the for loop. iterrows will give you the index of the row which we will store in our variable that we named index, and then the entire contents of the row will be stored in the value row as like a dictionary with each column name as the key(product_id,...) and the actual value of it. you can call it by referring to it by our variable name of row and putting column name as the key you are searching for in this dictionary. so it will be row["column_name"] and that will give you the value.
    mypdf.set_font(family="Times", size=10)
    mypdf.set_text_color(80,80,80)
    mypdf.cell(w=30, h=8, txt=str(row["product_id"]), border=1) #we put ln=1 after this if we want a line break and want curosr to start at next line. this will create a 30mm cell, write in it, and cursor will be at the end of this cell on the same line just 30mm away. so if we want something next to it, just define it below it will be next to it. ln will give you a line break of how many ever ypu tell it
    mypdf.cell(w=70, h=8, txt=str(row["product_name"]), border=1)#txt expects a string and row[value] rn is giing us numbers so we have to convert to string
    mypdf.cell(w=30, h=8, txt=str(row["amount_purchased"]), border=1)#
    mypdf.cell(w=30, h=8, txt=str(row["price_per_unit"]), border=1)
    mypdf.cell(w=30, h=8, txt=str(row["total_price"]), border=1, ln=1)#adding line break to last one so next row displays underneath, orginally was flowing off the page and wasn't visible

  #writning one final row to end of table with empty cells except total which will be grand total
  #this is outside of the writing rows for loop but still inside the giant one
  total = excelFile["total_price"].sum()#using the column name as a key on the whole excelFile dataframe gives us all the values of the column and sum just adds them up
  mypdf.set_font(family="Times", size=10, style='B')
  mypdf.set_text_color(80,80,80)
  mypdf.cell(w=30, h=8, txt="", border=1) 
  mypdf.cell(w=70, h=8, txt="", border=1)
  mypdf.cell(w=30, h=8, txt="", border=1)
  mypdf.cell(w=30, h=8, txt="", border=1)
  mypdf.cell(w=30, h=8, txt=str(total), border=1, ln=1)

  #writing underneath table
  mypdf.cell(w=30, h=8, txt=f"The total price to pay today is: {total}", ln=1)

  #company name and logo
  mypdf.set_font(family="Times", size=14, style='B')
  mypdf.cell(w=30, h=8, txt="PythonHow")
  mypdf.image("pythonhow.png", w=10)

  
  mypdf.output(f"reciepts/{filename}.pdf") #since filename is saved each loop, this allows us to grab the filename of the file we are looping over and make a pdf with the same name. have to add .pdf at the end
  